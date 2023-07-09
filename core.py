import os
import shutil
import datetime

from openpyxl import load_workbook

"""Запрос всех необходимых данных"""

work_directory = "/home/biglove/Desktop/Работа/НИЦ/2023/Биглов/"
new_business_trip = input("Введите город назначения: ")
started_at = input("Введите дату начала командировки: ")
ended_at = input("Введите дату окончания командировки: ")
transfer_expenses = int(input("Средства на транспорт: "))
representative_expenses = int(input("Представительские расходы: "))

"""
Создаем директории. Если город посещается первый раз - создается папка с названием города, а внутри папка с датами
командировки. Если визит повторный - создается в существующей папке директория с диапазоном дат.
"""

os.mkdir(work_directory + new_business_trip + f'/{started_at}-{ended_at}')
current_path = work_directory + new_business_trip + f'/{started_at}-{ended_at}'  # Директория по новой командировке

"""Копируем нужные файлы в новую директорию"""
os.rename(shutil.copy(work_directory + '/Для бухгалтерии.xlsx', current_path),
          current_path + f'/Для бухгалтерии-{new_business_trip}-{started_at}-{ended_at}.xlsx')

os.rename(shutil.copy(work_directory + '/Авансовый отчет.xlsx', current_path),
          current_path + f'/Авансовый отчет-{new_business_trip}-{started_at}-{ended_at}.xlsx')

os.rename(shutil.copy(work_directory + '/План-отчет.xlsx', current_path),
          current_path + f'/План-отчет-{new_business_trip}-{started_at}-{ended_at}.xlsx')

"""ВЫчисляем продолжительность командировки"""
start_date = datetime.datetime.strptime(started_at, '%d.%m.%Y')
end_date = datetime.datetime.strptime(ended_at, '%d.%m.%Y')
duration_of_business_trip = end_date - start_date

"""Открываем файл 'Для бухгалтерии' и пишем туда введенные данные"""
fn = f'{current_path}/Для бухгалтерии-{new_business_trip}-{started_at}-{ended_at}.xlsx'
wb = load_workbook(fn)
ws = wb['Лист1']

ws['D2'] = duration_of_business_trip.days + 1
ws['A2'] = f'{started_at}-{ended_at}'
ws['B2'] = new_business_trip
ws['E3'] = transfer_expenses
ws['E4'] = representative_expenses
wb.save(fn)
wb.close()


"""Открываем файл 'Авансовый отчет' и пишем туда введенные данные"""
fn1 = f'{current_path}/Авансовый отчет-{new_business_trip}-{started_at}-{ended_at}.xlsx'
wb1 = load_workbook(fn1)
ws1 = wb1['Лист1']

ws1['D2'] = 3
ws1['A2'] = f'{started_at}-{ended_at}'
ws1['B2'] = new_business_trip
ws1['E3'] = transfer_expenses
ws1['E4'] = representative_expenses
wb1.save(fn1)
wb1.close()
